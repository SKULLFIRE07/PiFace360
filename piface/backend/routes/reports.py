"""
PiFace Attendance System - Report Routes

GET  /daily/{date}          - Get or generate report data for a date
GET  /history               - List stored report files
GET  /download/{date}/excel - Download Excel report
GET  /download/{date}/pdf   - Download PDF report
POST /generate              - Trigger report generation in background
"""

import json
import logging
import os
import re
import threading
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from piface.backend.database import SessionLocal, get_db
from piface.backend.models import (
    AttendanceEvent,
    DailySummary,
    Person,
    SystemSetting,
)
from piface.backend.schemas import ApiResponse, ReportGenerateRequest
from piface.backend.security import require_auth

logger = logging.getLogger(__name__)

router = APIRouter()

REPORTS_DIR = Path("/opt/piface/reports")
DATE_REGEX = re.compile(r"^\d{4}-\d{2}-\d{2}$")

# Status colours used by both Excel and PDF generators
_STATUS_COLOURS = {
    "PRESENT": (0, 176, 80),       # green
    "ABSENT": (255, 0, 0),         # red
    "HALF_DAY": (255, 192, 0),     # yellow/amber
    "ON_LEAVE": (0, 112, 192),     # blue
    "HOLIDAY": (0, 112, 192),      # blue
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _validate_date(date_str: str) -> str:
    """Validate date format and return the string, or raise 400."""
    if not DATE_REGEX.match(date_str):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Expected YYYY-MM-DD.",
        )
    # Validate it parses to a real date
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date value.",
        )
    return date_str


def _safe_report_path(date_str: str, extension: str) -> Path:
    """Resolve a report file path and verify it is within REPORTS_DIR."""
    date_str = _validate_date(date_str)
    filename = f"report_{date_str}.{extension}"
    target = (REPORTS_DIR / date_str / filename).resolve()

    # Prevent path traversal
    if not str(target).startswith(str(REPORTS_DIR.resolve())):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid path.",
        )
    return target


def _get_company_name(db: Session) -> str:
    """Read company_name from system settings, with fallback."""
    row = (
        db.query(SystemSetting)
        .filter(SystemSetting.key == "company_name")
        .first()
    )
    return row.value if row and row.value else "PiFace Attendance System"


# ---------------------------------------------------------------------------
# Report Generation
# ---------------------------------------------------------------------------
def generate_report(date_str: str) -> None:
    """Generate Excel and PDF reports for the given date.

    This function creates its own database session so it can safely be called
    from a background thread.
    """
    from datetime import date as date_type

    db = SessionLocal()
    try:
        report_date = datetime.strptime(date_str, "%Y-%m-%d").date()

        # Query summaries joined with persons
        summaries = (
            db.query(DailySummary, Person)
            .join(Person, DailySummary.person_id == Person.id)
            .filter(DailySummary.date == report_date)
            .all()
        )

        # Query all events for the date
        events = (
            db.query(AttendanceEvent, Person)
            .join(Person, AttendanceEvent.person_id == Person.id)
            .filter(AttendanceEvent.date == report_date)
            .order_by(AttendanceEvent.timestamp)
            .all()
        )

        # Unknown persons seen that day
        unknown_events = (
            db.query(
                Person.id,
                Person.name,
                func.count(AttendanceEvent.id).label("event_count"),
                func.min(AttendanceEvent.timestamp).label("first_seen"),
                func.max(AttendanceEvent.timestamp).label("last_seen"),
            )
            .join(AttendanceEvent, AttendanceEvent.person_id == Person.id)
            .filter(
                AttendanceEvent.date == report_date,
                Person.is_unknown.is_(True),
            )
            .group_by(Person.id)
            .all()
        )

        company_name = _get_company_name(db)

        # Ensure output directory
        out_dir = REPORTS_DIR / date_str
        out_dir.mkdir(parents=True, exist_ok=True)

        _generate_excel(
            out_dir / f"report_{date_str}.xlsx",
            date_str,
            company_name,
            summaries,
            events,
            unknown_events,
        )
        _generate_pdf(
            out_dir / f"report_{date_str}.pdf",
            date_str,
            company_name,
            summaries,
            unknown_events,
        )

        logger.info("Reports generated for %s", date_str)
    except Exception:
        logger.exception("Failed to generate reports for %s", date_str)
    finally:
        db.close()


def _generate_excel(
    path: Path,
    date_str: str,
    company_name: str,
    summaries: list,
    events: list,
    unknown_events: list,
) -> None:
    """Generate a styled Excel workbook with three sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # -- Sheet 1: Summary -------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Company header
    ws_summary.merge_cells("A1:H1")
    header_cell = ws_summary["A1"]
    header_cell.value = f"{company_name} - Attendance Report - {date_str}"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal="center")

    columns = [
        "Name",
        "Employee ID",
        "Department",
        "First In",
        "Last Out",
        "Breaks (min)",
        "Total Hours",
        "Status",
    ]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws_summary.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for row_idx, (summary, person) in enumerate(summaries, start=4):
        ws_summary.cell(row=row_idx, column=1, value=person.name)
        ws_summary.cell(row=row_idx, column=2, value=person.employee_id or "")
        ws_summary.cell(row=row_idx, column=3, value=person.department or "")
        ws_summary.cell(
            row=row_idx,
            column=4,
            value=summary.first_in_time.strftime("%H:%M:%S") if summary.first_in_time else "",
        )
        ws_summary.cell(
            row=row_idx,
            column=5,
            value=summary.last_out_time.strftime("%H:%M:%S") if summary.last_out_time else "",
        )
        ws_summary.cell(row=row_idx, column=6, value=summary.total_break_minutes or 0)
        ws_summary.cell(
            row=row_idx,
            column=7,
            value=round(summary.total_hours_worked, 2) if summary.total_hours_worked else 0,
        )

        status_val = summary.status or "ABSENT"
        status_cell = ws_summary.cell(row=row_idx, column=8, value=status_val)

        rgb = _STATUS_COLOURS.get(status_val)
        if rgb:
            hex_color = "{:02X}{:02X}{:02X}".format(*rgb)
            status_cell.fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            # Use white text for dark backgrounds
            if status_val in ("PRESENT", "ABSENT", "ON_LEAVE", "HOLIDAY"):
                status_cell.font = Font(color="FFFFFF", bold=True)
            else:
                status_cell.font = Font(bold=True)

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(columns) + 1):
        ws_summary.column_dimensions[
            ws_summary.cell(row=3, column=col_idx).column_letter
        ].width = 16

    # -- Sheet 2: Event Log ------------------------------------------------
    ws_events = wb.create_sheet("Event Log")

    event_columns = [
        "Person",
        "Employee ID",
        "Event Type",
        "Timestamp",
        "Confidence",
        "Manual",
    ]
    for col_idx, col_name in enumerate(event_columns, start=1):
        cell = ws_events.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for row_idx, (event, person) in enumerate(events, start=2):
        ws_events.cell(row=row_idx, column=1, value=person.name)
        ws_events.cell(row=row_idx, column=2, value=person.employee_id or "")
        ws_events.cell(row=row_idx, column=3, value=event.event_type)
        ws_events.cell(
            row=row_idx,
            column=4,
            value=event.timestamp.strftime("%Y-%m-%d %H:%M:%S") if event.timestamp else "",
        )
        ws_events.cell(
            row=row_idx,
            column=5,
            value=round(event.confidence, 3) if event.confidence else "",
        )
        ws_events.cell(row=row_idx, column=6, value="Yes" if event.is_manual else "No")

    for col_idx in range(1, len(event_columns) + 1):
        ws_events.column_dimensions[
            ws_events.cell(row=1, column=col_idx).column_letter
        ].width = 18

    # -- Sheet 3: Unknowns -------------------------------------------------
    ws_unknowns = wb.create_sheet("Unknowns")

    unknown_columns = ["ID", "Name", "Events", "First Seen", "Last Seen"]
    for col_idx, col_name in enumerate(unknown_columns, start=1):
        cell = ws_unknowns.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for row_idx, row in enumerate(unknown_events, start=2):
        ws_unknowns.cell(row=row_idx, column=1, value=row.id)
        ws_unknowns.cell(row=row_idx, column=2, value=row.name)
        ws_unknowns.cell(row=row_idx, column=3, value=row.event_count)
        ws_unknowns.cell(
            row=row_idx,
            column=4,
            value=row.first_seen.strftime("%H:%M:%S") if row.first_seen else "",
        )
        ws_unknowns.cell(
            row=row_idx,
            column=5,
            value=row.last_seen.strftime("%H:%M:%S") if row.last_seen else "",
        )

    for col_idx in range(1, len(unknown_columns) + 1):
        ws_unknowns.column_dimensions[
            ws_unknowns.cell(row=1, column=col_idx).column_letter
        ].width = 16

    wb.save(str(path))


def _generate_pdf(
    path: Path,
    date_str: str,
    company_name: str,
    summaries: list,
    unknown_events: list,
) -> None:
    """Generate a PDF report with summary table, colour coding, and footer."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    page_width, page_height = landscape(A4)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # -- Header ------------------------------------------------------------
    elements.append(Paragraph(company_name, styles["Title"]))
    elements.append(
        Paragraph(f"Daily Attendance Report - {date_str}", styles["Heading2"])
    )
    elements.append(Spacer(1, 10 * mm))

    # -- Summary table -----------------------------------------------------
    table_headers = [
        "Name",
        "Employee ID",
        "Dept",
        "First In",
        "Last Out",
        "Breaks (min)",
        "Total Hours",
        "Status",
    ]
    table_data = [table_headers]

    total_present = 0
    total_absent = 0
    total_hours_list: list[float] = []

    rl_status_colours = {
        "PRESENT": colors.Color(0 / 255, 176 / 255, 80 / 255),
        "ABSENT": colors.Color(255 / 255, 0 / 255, 0 / 255),
        "HALF_DAY": colors.Color(255 / 255, 192 / 255, 0 / 255),
        "ON_LEAVE": colors.Color(0 / 255, 112 / 255, 192 / 255),
        "HOLIDAY": colors.Color(0 / 255, 112 / 255, 192 / 255),
    }

    for summary, person in summaries:
        status_val = summary.status or "ABSENT"
        hours = round(summary.total_hours_worked, 2) if summary.total_hours_worked else 0

        if status_val == "PRESENT":
            total_present += 1
        elif status_val == "ABSENT":
            total_absent += 1

        if summary.total_hours_worked:
            total_hours_list.append(summary.total_hours_worked)

        table_data.append([
            person.name,
            person.employee_id or "",
            person.department or "",
            summary.first_in_time.strftime("%H:%M:%S") if summary.first_in_time else "",
            summary.last_out_time.strftime("%H:%M:%S") if summary.last_out_time else "",
            str(summary.total_break_minutes or 0),
            str(hours),
            status_val,
        ])

    if len(table_data) > 1:
        col_widths = [
            page_width * 0.15,
            page_width * 0.10,
            page_width * 0.10,
            page_width * 0.10,
            page_width * 0.10,
            page_width * 0.10,
            page_width * 0.10,
            page_width * 0.10,
        ]
        table = Table(table_data, colWidths=col_widths)

        table_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]

        # Colour-code the status column
        for row_idx in range(1, len(table_data)):
            status_val = table_data[row_idx][7]
            colour = rl_status_colours.get(status_val)
            if colour:
                table_style.append(("BACKGROUND", (7, row_idx), (7, row_idx), colour))
                table_style.append(("TEXTCOLOR", (7, row_idx), (7, row_idx), colors.white))

        table.setStyle(TableStyle(table_style))
        elements.append(table)
    else:
        elements.append(Paragraph("No attendance data for this date.", styles["Normal"]))

    # -- Footer stats ------------------------------------------------------
    elements.append(Spacer(1, 10 * mm))
    avg_hours = round(sum(total_hours_list) / len(total_hours_list), 2) if total_hours_list else 0
    footer_text = (
        f"Total Present: {total_present}  |  "
        f"Total Absent: {total_absent}  |  "
        f"Average Hours: {avg_hours}"
    )
    elements.append(Paragraph(footer_text, styles["Normal"]))

    # -- Build with page numbers -------------------------------------------
    def _add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(
            page_width - 15 * mm,
            10 * mm,
            f"Page {doc.page}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_add_page_number, onLaterPages=_add_page_number)


# ---------------------------------------------------------------------------
# GET /daily/{date} - Get report data for a date
# ---------------------------------------------------------------------------
@router.get("/daily/{date}", response_model=ApiResponse)
def get_daily_report(
    date: str,
    db: Session = Depends(get_db),
    current_user: str = Depends(require_auth),
) -> ApiResponse:
    """Return daily summary data for all persons on the given date."""
    date_str = _validate_date(date)
    report_date = datetime.strptime(date_str, "%Y-%m-%d").date()

    summaries = (
        db.query(DailySummary, Person)
        .join(Person, DailySummary.person_id == Person.id)
        .filter(DailySummary.date == report_date)
        .all()
    )

    data = []
    for summary, person in summaries:
        data.append(
            {
                "person_id": person.id,
                "name": person.name,
                "employee_id": person.employee_id,
                "department": person.department,
                "date": date_str,
                "first_in_time": (
                    summary.first_in_time.isoformat() if summary.first_in_time else None
                ),
                "last_out_time": (
                    summary.last_out_time.isoformat() if summary.last_out_time else None
                ),
                "total_hours_worked": summary.total_hours_worked,
                "total_break_minutes": summary.total_break_minutes,
                "status": summary.status,
                "is_late": summary.is_late,
                "is_early_leave": summary.is_early_leave,
                "overtime_minutes": summary.overtime_minutes,
            }
        )

    return ApiResponse(success=True, data=data)


# ---------------------------------------------------------------------------
# GET /history - List stored report files
# ---------------------------------------------------------------------------
@router.get("/history", response_model=ApiResponse)
def list_report_history(
    current_user: str = Depends(require_auth),
) -> ApiResponse:
    """Scan the reports directory and return available report files."""
    if not REPORTS_DIR.exists():
        return ApiResponse(success=True, data=[])

    reports = []
    for entry in sorted(REPORTS_DIR.iterdir()):
        if not entry.is_dir() or not DATE_REGEX.match(entry.name):
            continue

        date_str = entry.name
        excel_path = entry / f"report_{date_str}.xlsx"
        pdf_path = entry / f"report_{date_str}.pdf"

        # Use the most recent modification time as generated_at
        generated_at = None
        for fpath in (excel_path, pdf_path):
            if fpath.exists():
                mtime = datetime.fromtimestamp(
                    fpath.stat().st_mtime, tz=timezone.utc
                )
                if generated_at is None or mtime > generated_at:
                    generated_at = mtime

        reports.append(
            {
                "date": date_str,
                "has_excel": excel_path.exists(),
                "has_pdf": pdf_path.exists(),
                "generated_at": generated_at.isoformat() if generated_at else None,
            }
        )

    return ApiResponse(success=True, data=reports)


# ---------------------------------------------------------------------------
# GET /download/{date}/excel
# ---------------------------------------------------------------------------
@router.get("/download/{date}/excel")
def download_excel(
    date: str,
    current_user: str = Depends(require_auth),
):
    """Download the Excel report for the given date."""
    path = _safe_report_path(date, "xlsx")
    if not path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Excel report for {date} not found",
        )
    return FileResponse(
        path=str(path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name,
    )


# ---------------------------------------------------------------------------
# GET /download/{date}/pdf
# ---------------------------------------------------------------------------
@router.get("/download/{date}/pdf")
def download_pdf(
    date: str,
    current_user: str = Depends(require_auth),
):
    """Download the PDF report for the given date."""
    path = _safe_report_path(date, "pdf")
    if not path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"PDF report for {date} not found",
        )
    return FileResponse(
        path=str(path),
        media_type="application/pdf",
        filename=path.name,
    )


# ---------------------------------------------------------------------------
# POST /generate - Trigger report generation
# ---------------------------------------------------------------------------
@router.post("/generate", response_model=ApiResponse, status_code=status.HTTP_202_ACCEPTED)
def trigger_report_generation(
    body: ReportGenerateRequest,
    current_user: str = Depends(require_auth),
) -> ApiResponse:
    """Manually trigger report generation for a date. Runs in a background thread."""
    date_str = _validate_date(body.date)

    thread = threading.Thread(
        target=generate_report,
        args=(date_str,),
        name=f"report-gen-{date_str}",
        daemon=True,
    )
    thread.start()

    logger.info("User %s triggered report generation for %s", current_user, date_str)

    return ApiResponse(
        success=True,
        data={"date": date_str, "status": "generation_started"},
    )


# ---------------------------------------------------------------------------
# GET / - Date-range report (frontend calls GET /reports?date_from=...&date_to=...)
# ---------------------------------------------------------------------------
@router.get("/", response_model=ApiResponse)
def get_report_range(
    date_from: str = Query(default=None),
    date_to: str = Query(default=None),
    db: Session = Depends(get_db),
    current_user: str = Depends(require_auth),
) -> ApiResponse:
    """Return attendance summary for a date range."""
    from datetime import date as date_type

    try:
        start = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else date_type.today()
        end = datetime.strptime(date_to, "%Y-%m-%d").date() if date_to else start
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    summaries = (
        db.query(DailySummary, Person)
        .join(Person, DailySummary.person_id == Person.id)
        .filter(DailySummary.date >= start, DailySummary.date <= end)
        .order_by(DailySummary.date.desc())
        .all()
    )

    # Per-employee aggregated summary
    from collections import defaultdict
    emp_agg = defaultdict(lambda: {
        "days_present": 0, "days_absent": 0, "days_leave": 0,
        "total_hours": 0.0, "late_count": 0,
    })

    daily_counts = defaultdict(lambda: {"present": 0, "absent": 0, "leave": 0})
    breakdown = {"present": 0, "absent": 0, "leave": 0}

    records = []
    for summary, person in summaries:
        records.append({
            "person_id": person.id,
            "name": person.name,
            "employee_id": person.employee_id,
            "department": person.department,
            "date": summary.date.isoformat() if summary.date else None,
            "first_in_time": summary.first_in_time.isoformat() if summary.first_in_time else None,
            "last_out_time": summary.last_out_time.isoformat() if summary.last_out_time else None,
            "total_hours_worked": summary.total_hours_worked,
            "status": summary.status,
            "is_late": summary.is_late,
        })

        agg = emp_agg[person.id]
        agg["name"] = person.name
        agg["employee_id"] = person.employee_id
        agg["department"] = person.department

        st = (summary.status or "").upper()
        day_key = summary.date.isoformat() if summary.date else "unknown"

        if st in ("PRESENT", "IN", "OUT"):
            agg["days_present"] += 1
            daily_counts[day_key]["present"] += 1
            breakdown["present"] += 1
        elif st in ("ABSENT", "NOT_ARRIVED"):
            agg["days_absent"] += 1
            daily_counts[day_key]["absent"] += 1
            breakdown["absent"] += 1
        elif st in ("ON_LEAVE", "LEAVE"):
            agg["days_leave"] += 1
            daily_counts[day_key]["leave"] += 1
            breakdown["leave"] += 1

        agg["total_hours"] += summary.total_hours_worked or 0.0
        if summary.is_late:
            agg["late_count"] += 1

    # Build summary rows (per-employee)
    summary_rows = []
    for pid, agg in emp_agg.items():
        summary_rows.append({
            "person_id": pid,
            "name": agg["name"],
            "employee_id": agg["employee_id"],
            "department": agg["department"],
            "days_present": agg["days_present"],
            "days_absent": agg["days_absent"],
            "days_leave": agg["days_leave"],
            "total_hours": round(agg["total_hours"], 2),
            "late_count": agg["late_count"],
        })

    # Build daily chart data
    chart_data = [
        {"date": d, **counts}
        for d, counts in sorted(daily_counts.items())
    ]

    return ApiResponse(success=True, data={
        "summary": summary_rows,
        "chart_data": chart_data,
        "breakdown": breakdown,
        "records": records,
    })


# ---------------------------------------------------------------------------
# GET /download/excel - Download Excel report (query-param style)
# ---------------------------------------------------------------------------
@router.get("/download/excel")
def download_excel_range(
    date_from: str = Query(default=None),
    date_to: str = Query(default=None),
    current_user: str = Depends(require_auth),
):
    """Download Excel report. Falls back to date_from as single-date download."""
    date_str = date_from or date_to
    if not date_str:
        raise HTTPException(status_code=400, detail="date_from is required")
    return download_excel(date=date_str, current_user=current_user)


# ---------------------------------------------------------------------------
# GET /download/pdf - Download PDF report (query-param style)
# ---------------------------------------------------------------------------
@router.get("/download/pdf")
def download_pdf_range(
    date_from: str = Query(default=None),
    date_to: str = Query(default=None),
    current_user: str = Depends(require_auth),
):
    """Download PDF report. Falls back to date_from as single-date download."""
    date_str = date_from or date_to
    if not date_str:
        raise HTTPException(status_code=400, detail="date_from is required")
    return download_pdf(date=date_str, current_user=current_user)
